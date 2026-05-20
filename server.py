#!/usr/bin/env python3
"""
the 100's — Servidor Python + MySQL
Sistema de Gestão de Loja de Vinhos Premium

Projeto pessoal · © 2026 Pedro Gabriel Matias Gomes
Bottled Memories — www.the-100s.com

Corre: python3 server.py
Site:  http://localhost:8080
"""

from flask import Flask, request, jsonify, send_from_directory, g
from flask_cors import CORS
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
import datetime
import json
import logging
import os
import random
import secrets
import threading

# ── Logging ───────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%H:%M:%S'
)
log = logging.getLogger('vinhadouro')

# ── .env loader (sem dependência externa) ────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

def _load_dotenv():
    """Carrega .env se existir, sem sobrepor variáveis já definidas."""
    path = os.path.join(BASE_DIR, '.env')
    if not os.path.exists(path):
        return
    try:
        with open(path, 'r', encoding='utf-8') as f:
            for raw in f:
                line = raw.strip()
                if not line or line.startswith('#') or '=' not in line:
                    continue
                k, _, v = line.partition('=')
                k = k.strip()
                v = v.strip().strip('"').strip("'")
                if k and k not in os.environ:
                    os.environ[k] = v
    except Exception as e:
        log.warning(".env não pôde ser lido: %s", e)

_load_dotenv()

# ── MySQL connection ──────────────────────────────────────
try:
    import pymysql
    import pymysql.cursors
    MYSQL_AVAILABLE = True
except ImportError:
    MYSQL_AVAILABLE = False

import sqlite3  # fallback

# Onde estão os ficheiros estáticos (HTML/CSS/JS/assets). Movidos para frontend/ na v9.2.
FRONTEND_DIR = os.path.join(BASE_DIR, 'frontend')

app = Flask(__name__, static_folder=FRONTEND_DIR, static_url_path='')

# CORS — restringido a origens conhecidas (configurável via .env)
_cors_origins = [o.strip() for o in os.getenv(
    'CORS_ORIGINS',
    'http://localhost:8080,http://127.0.0.1:8080'
).split(',') if o.strip()]
CORS(app, resources={r"/api/*": {"origins": _cors_origins}}, supports_credentials=False)


@app.after_request
def _security_headers(response):
    response.headers.setdefault('X-Content-Type-Options', 'nosniff')
    response.headers.setdefault('X-Frame-Options', 'SAMEORIGIN')
    response.headers.setdefault('Referrer-Policy', 'strict-origin-when-cross-origin')
    return response

# ── DB Config (a partir de .env / variáveis de ambiente) ─
DB_TYPE   = 'mysql' if MYSQL_AVAILABLE else 'sqlite'
MYSQL_CFG = {
    'host':     os.getenv('MYSQL_HOST', 'localhost'),
    'port':     int(os.getenv('MYSQL_PORT', '3306')),
    'user':     os.getenv('MYSQL_USER', 'root'),
    'password': os.getenv('MYSQL_PASSWORD', '2006'),
    'database': os.getenv('MYSQL_DATABASE', 'vinhadouro'),
    'charset':  'utf8mb4',
    'cursorclass': None,  # set at runtime
    'autocommit': True,
}

# ── Auth: tokens de sessão em memória ────────────────────
# Mapeia token -> {'user_id', 'username', 'role', 'nome', 'expires_at'}
# Em produção real isto seria Redis/JWT. Para projeto académico, em memória chega.
_SESSIONS = {}
_SESSIONS_LOCK = threading.Lock()
SESSION_TTL = datetime.timedelta(hours=int(os.getenv('SESSION_TTL_HOURS', '8')))

def _new_token(user):
    token = secrets.token_urlsafe(32)
    expires = datetime.datetime.utcnow() + SESSION_TTL
    with _SESSIONS_LOCK:
        _SESSIONS[token] = {
            'user_id':  user['id'],
            'username': user['username'],
            'role':     user['role'],
            'nome':     user.get('nome') or user['username'],
            'expires_at': expires,
        }
    return token, expires

def _resolve_token(token):
    if not token:
        return None
    with _SESSIONS_LOCK:
        sess = _SESSIONS.get(token)
        if not sess:
            return None
        if sess['expires_at'] < datetime.datetime.utcnow():
            _SESSIONS.pop(token, None)
            return None
        return sess

def _revoke_token(token):
    with _SESSIONS_LOCK:
        _SESSIONS.pop(token, None)

def _get_request_token():
    auth = request.headers.get('Authorization', '')
    if auth.startswith('Bearer '):
        return auth[7:].strip()
    return None

def requires_auth(fn):
    """Exige que o pedido traga um token válido."""
    @wraps(fn)
    def wrapper(*args, **kwargs):
        sess = _resolve_token(_get_request_token())
        if not sess:
            return jsonify({'error': 'Não autenticado'}), 401
        g.user = sess
        return fn(*args, **kwargs)
    return wrapper

def requires_role(*roles):
    """Exige autenticação e role permitido."""
    allowed = set(roles)
    def deco(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            sess = _resolve_token(_get_request_token())
            if not sess:
                return jsonify({'error': 'Não autenticado'}), 401
            if sess['role'] not in allowed:
                return jsonify({'error': 'Sem permissão para esta operação'}), 403
            g.user = sess
            return fn(*args, **kwargs)
        return wrapper
    return deco

def get_db():
    """Obter conexão com a base de dados (MySQL ou SQLite)."""
    if DB_TYPE == 'mysql' and MYSQL_AVAILABLE:
        try:
            cfg = {**MYSQL_CFG, 'cursorclass': pymysql.cursors.DictCursor}
            conn = pymysql.connect(**cfg)
            return conn, 'mysql'
        except Exception as e:
            log.warning("MySQL falhou (%s), a usar SQLite como fallback", e)
    path = os.path.join(BASE_DIR, 'vinhadouro.db')
    conn = sqlite3.connect(path)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn, 'sqlite'

def db_exec(conn, db_type, sql, params=()):
    """Executar uma query, tratando diferenças MySQL vs SQLite."""
    if db_type == 'mysql':
        sql = sql.replace('?', '%s')
        with conn.cursor() as cur:
            cur.execute(sql, params)
            return cur
    else:
        return conn.execute(sql, params)

def db_fetchall(conn, db_type, sql, params=()):
    """Buscar todos os registos."""
    cur = db_exec(conn, db_type, sql, params)
    if db_type == 'mysql':
        return cur.fetchall()
    return [dict(r) for r in cur.fetchall()]

def db_fetchone(conn, db_type, sql, params=()):
    """Buscar um único registo."""
    cur = db_exec(conn, db_type, sql, params)
    if db_type == 'mysql':
        return cur.fetchone()
    r = cur.fetchone()
    return dict(r) if r else None

# ── Init DB ───────────────────────────────────────────────
def init_db():
    """Inicializar base de dados"""
    conn, db_type = get_db()

    if db_type == 'mysql':
        _init_mysql(conn)
    else:
        _init_sqlite(conn)

    conn.close()
    log.info("Base de dados inicializada (%s)", db_type.upper())

def _init_mysql(conn):
    """Inicializar tabelas MySQL básicas (compatibilidade)"""
    with conn.cursor() as c:
        c.execute("""
        CREATE TABLE IF NOT EXISTS utilizadores (
            id INT AUTO_INCREMENT PRIMARY KEY,
            pessoa_id INT,
            username VARCHAR(100) UNIQUE NOT NULL,
            password_hash VARCHAR(255) NOT NULL,
            role ENUM('FUNCIONARIO','GERENTE','ARMAZENISTA','ADMIN') DEFAULT 'FUNCIONARIO',
            ativo TINYINT(1) DEFAULT 1,
            criado_em DATETIME DEFAULT CURRENT_TIMESTAMP
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS pessoas (
            id INT AUTO_INCREMENT PRIMARY KEY,
            nome VARCHAR(150) NOT NULL,
            email VARCHAR(150),
            telefone VARCHAR(30)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS funcionarios (
            id INT AUTO_INCREMENT PRIMARY KEY,
            pessoa_id INT REFERENCES pessoas(id),
            cargo VARCHAR(100),
            salario DECIMAL(10,2) DEFAULT 0,
            ativo TINYINT(1) DEFAULT 1,
            nivel VARCHAR(30) DEFAULT 'FUNCIONARIO',
            criado_em DATETIME DEFAULT CURRENT_TIMESTAMP
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS vinhos (
            id INT AUTO_INCREMENT PRIMARY KEY,
            nome VARCHAR(200) NOT NULL,
            tipo VARCHAR(50),
            regiao VARCHAR(100),
            produtor VARCHAR(150),
            ano_colheita INT,
            preco DECIMAL(10,2) DEFAULT 0,
            quantidade INT DEFAULT 0,
            descricao TEXT,
            imagem_url VARCHAR(500)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS vendas (
            id INT AUTO_INCREMENT PRIMARY KEY,
            codigo VARCHAR(50),
            pessoa_id INT,
            cliente_id INT,
            funcionario_id INT,
            total DECIMAL(10,2) DEFAULT 0,
            metodo_pagamento ENUM('DINHEIRO','CARTAO','MB_WAY','TRANSFERENCIA'),
            estado ENUM('PENDENTE','CONCLUIDA','CANCELADA') DEFAULT 'CONCLUIDA',
            notas TEXT,
            criado_em DATETIME DEFAULT CURRENT_TIMESTAMP
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS itens_venda (
            id INT AUTO_INCREMENT PRIMARY KEY,
            venda_id INT REFERENCES vendas(id),
            vinho_id INT REFERENCES vinhos(id),
            quantidade INT DEFAULT 1,
            preco_unit DECIMAL(10,2) DEFAULT 0,
            subtotal DECIMAL(10,2) DEFAULT 0
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS caves (
            id INT AUTO_INCREMENT PRIMARY KEY,
            nome VARCHAR(200),
            localizacao VARCHAR(300),
            capacidade INT DEFAULT 0,
            temperatura_ideal DECIMAL(4,1),
            humidade_ideal DECIMAL(4,1),
            notas TEXT,
            ativo TINYINT DEFAULT 1,
            criado_em DATETIME DEFAULT CURRENT_TIMESTAMP
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS movimentos_stock (
            id INT AUTO_INCREMENT PRIMARY KEY,
            vinho_id INT,
            tipo ENUM('ENTRADA','SAIDA','AJUSTE','TRANSFERENCIA'),
            quantidade INT,
            motivo VARCHAR(300),
            funcionario_id INT,
            criado_em DATETIME DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (vinho_id) REFERENCES vinhos(id)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS clientes (
            id INT AUTO_INCREMENT PRIMARY KEY,
            pessoa_id INT,
            nif VARCHAR(20),
            data_nascimento DATE,
            preferencias JSON,
            ativo TINYINT DEFAULT 1,
            criado_em DATETIME DEFAULT CURRENT_TIMESTAMP
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)

        conn.commit()

        # Seed data
        c.execute("SELECT COUNT(*) as cnt FROM utilizadores")
        if c.fetchone()['cnt'] == 0:
            _seed_mysql(conn, c)

def _seed_mysql(conn, c):
    """Preencher dados iniciais MySQL"""
    # Inserir pessoas primeiro
    pessoas = [
        ("António Ferreira", "gerente@vinhadouro.pt", "912000001"),
        ("Sofia Martins", "loja@vinhadouro.pt", "912000002"),
        ("João Rodrigues", "stock@vinhadouro.pt", "912000003"),
        ("Admin Sistema", "admin@vinhadouro.pt", "912000004"),
    ]
    for p in pessoas:
        c.execute("INSERT IGNORE INTO pessoas (nome,email,telefone) VALUES (%s,%s,%s)", p)
    conn.commit()
    utilizadores = [
        (1, "gerente", generate_password_hash("1234"), "GERENTE"),
        (2, "loja",    generate_password_hash("1234"), "FUNCIONARIO"),
        (3, "stock",   generate_password_hash("1234"), "ARMAZENISTA"),
        (4, "admin",   generate_password_hash("1234"), "ADMIN"),
    ]
    for u in utilizadores:
        c.execute("INSERT IGNORE INTO utilizadores (pessoa_id,username,password_hash,role) VALUES (%s,%s,%s,%s)", u)

    # Catálogo curado — 24 vinhos cobrindo Tinto, Branco, Rosé, Espumante, Porto e Madeira.
    vinhos = [
        # ── TINTOS ──
        ("Barca Velha","Tinto","Douro","Casa Ferreirinha",2011,280.00,3,"O ícone do Douro. Complexo, elegante, longevidade incomparável.",None),
        ("Quinta do Crasto Reserva","Tinto","Douro","Quinta do Crasto",2019,28.50,28,"Frutos vermelhos, taninos sedosos, clássico do Douro Superior.",None),
        ("Niepoort Redoma Tinto","Tinto","Douro","Niepoort",2020,27.00,15,"Elegante, terroir duriense puro, evolução em garrafa.",None),
        ("Mouchão Tonel 3-4","Tinto","Alentejo","Herdade do Mouchão",2015,62.00,5,"Castelão histórico do Alentejo. Estrutura imponente.",None),
        ("Esporão Reserva","Tinto","Alentejo","Herdade do Esporão",2019,18.90,40,"Aragonez e Trincadeira em harmonia. Madeira fina e fruta madura.",None),
        ("Quinta da Pellada Primus","Tinto","Dão","Quinta da Pellada",2018,32.00,11,"Touriga Nacional de vinhas velhas. Premium do Dão.",None),
        ("Luís Pato Vinhas Velhas","Tinto","Bairrada","Luís Pato",2018,22.00,16,"Baga clássica, acidez vibrante, taninos firmes.",None),
        # ── BRANCOS ──
        ("Anselmo Mendes Alvarinho","Branco","Vinho Verde","Anselmo Mendes",2022,16.50,30,"Frescura mineral, notas cítricas e floral. Alvarinho de eleição.",None),
        ("Niepoort Redoma Branco","Branco","Douro","Niepoort",2021,24.50,18,"Branco de altitude. Mineralidade e potencial de guarda.",None),
        ("Esporão Branco Reserva","Branco","Alentejo","Herdade do Esporão",2022,14.80,35,"Tropical com final mineral. Acompanha peixe e marisco.",None),
        ("Casa de Santar Reserva","Branco","Dão","Casa de Santar",2021,12.90,28,"Encruzado equilibrado. Pêssego, pêra e amêndoa.",None),
        # ── ROSÉ ──
        ("Quinta dos Murças Rosé","Rosé","Douro","Quinta dos Murças",2023,12.50,32,"Rosé do Douro de cor salmão. Morango e framboesa.",None),
        ("Mateus Rosé","Rosé","Bairrada","Sogrape",2023,5.90,80,"Ícone português. Fresco, ligeiramente frisante, clássico.",None),
        # ── ESPUMANTES ──
        ("Raposeira Super Reserva","Espumante","Bairrada","Raposeira",2018,12.40,40,"Bagas finas, elegante na boca, acabamento seco impecável.",None),
        ("Murganheira Bruto","Espumante","Távora-Varosa","Murganheira",2019,15.50,22,"Método clássico. Cremosidade e frescura granítica.",None),
        # ── PORTO ──
        ("Graham's 20 Years Old Tawny","Porto","Douro","Graham's",None,68.00,12,"Tawny envelhecido. Frutos secos, especiarias, mel.",None),
        ("Taylor's Vintage 2017","Porto","Douro","Taylor's",2017,135.00,6,"Vintage histórico. Concentração, taninos e doçura em harmonia.",None),
        ("Niepoort LBV","Porto","Douro","Niepoort",2018,22.50,24,"Late Bottled Vintage. Fruta preta e estrutura no copo.",None),
        ("Ferreira Dona Antónia Reserva","Porto","Douro","Sogevinus",None,18.90,30,"Tawny Reserva. Clássico, equilibrado, intemporal.",None),
        ("Quinta do Noval LBV","Porto","Douro","Quinta do Noval",2017,28.00,16,"Vinhedo singular. Concentração e elegância.",None),
        # ── MADEIRA ──
        ("Blandy's 10 Years Malmsey","Madeira","Madeira","Blandy's",None,45.00,10,"Doce, complexo. Mel, especiarias, café tostado.",None),
        ("Henriques & Henriques 15 Anos Sercial","Madeira","Madeira","Henriques & Henriques",None,55.00,8,"Seco, frutos secos, acidez vibrante. O Madeira aristocrático.",None),
        ("Barbeito Verdelho 5 Anos","Madeira","Madeira","Barbeito",None,28.00,14,"Meio-seco. Caramelo salgado e nozes. Versátil à mesa.",None),
        ("D'Oliveiras Boal 1989","Madeira","Madeira","D'Oliveiras",1989,180.00,3,"Colheita raríssima. Doce, untuoso, eternamente jovem.",None),
    ]
    for v in vinhos:
        c.execute("""INSERT IGNORE INTO vinhos (nome,tipo,regiao,produtor,ano_colheita,preco,quantidade,descricao,imagem_url)
                     VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)""", v)

    pessoas = [
        ("António Ferreira","antonio@vinhadouro.pt","912345678"),
        ("Sofia Martins","sofia@vinhadouro.pt","913456789"),
        ("João Rodrigues","joao@vinhadouro.pt","914567890"),
        ("Maria Costa","maria@vinhadouro.pt","915678901"),
    ]
    ids = []
    for p in pessoas:
        c.execute("INSERT IGNORE INTO pessoas (nome,email,telefone) VALUES (%s,%s,%s)", p)
        ids.append(c.lastrowid)

    funcionarios = [
        (ids[0],"Gerente",2800.00,1,"GERENTE"),
        (ids[1],"Operador POS",1200.00,1,"FUNCIONARIO"),
        (ids[2],"Armazenista",1150.00,1,"ARMAZENISTA"),
        (ids[3],"Operador POS",1200.00,1,"FUNCIONARIO"),
    ]
    for f in funcionarios:
        c.execute("INSERT IGNORE INTO funcionarios (pessoa_id,cargo,salario,ativo,nivel) VALUES (%s,%s,%s,%s,%s)", f)

    metodos = ["CARTAO","MB_WAY","DINHEIRO","CARTAO"]
    for i in range(8):
        dias = random.randint(0,30)
        dt = (datetime.datetime.now() - datetime.timedelta(days=dias)).strftime("%Y-%m-%d %H:%M:%S")
        total = round(random.uniform(15,200),2)
        c.execute("INSERT INTO vendas (funcionario_id,criado_em,metodo_pagamento,estado,total) VALUES (%s,%s,%s,%s,%s)",
                  (1, dt, metodos[i%4], "CONCLUIDA", total))
        vid = c.lastrowid
        # Distribuir vendas por todos os tipos (24 vinhos no seed)
        c.execute("INSERT INTO itens_venda (venda_id,vinho_id,quantidade,preco_unit,subtotal) VALUES (%s,%s,%s,%s,%s)",
                  (vid, (i % 24) + 1, 1, total, total))

    conn.commit()

def _init_sqlite(conn):
    """Inicializar tabelas SQLite básicas"""
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS pessoas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nome TEXT NOT NULL,
        email TEXT,
        telefone TEXT
    );
    CREATE TABLE IF NOT EXISTS utilizadores (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        pessoa_id INTEGER,
        username TEXT NOT NULL UNIQUE,
        password_hash TEXT NOT NULL,
        role TEXT DEFAULT 'FUNCIONARIO',
        ativo INTEGER DEFAULT 1,
        criado_em TEXT DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (pessoa_id) REFERENCES pessoas(id)
    );
    CREATE TABLE IF NOT EXISTS funcionarios (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        pessoa_id INTEGER, cargo TEXT, salario REAL DEFAULT 0,
        ativo INTEGER DEFAULT 1, nivel TEXT DEFAULT 'FUNCIONARIO',
        criado_em TEXT DEFAULT CURRENT_TIMESTAMP
    );
    CREATE TABLE IF NOT EXISTS vinhos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nome TEXT NOT NULL, tipo TEXT, regiao TEXT, produtor TEXT,
        ano_colheita INTEGER, preco REAL DEFAULT 0, quantidade INTEGER DEFAULT 0,
        descricao TEXT, imagem_url TEXT
    );
    CREATE TABLE IF NOT EXISTS vendas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        codigo TEXT,
        pessoa_id INTEGER,
        cliente_id INTEGER,
        funcionario_id INTEGER,
        total REAL DEFAULT 0,
        metodo_pagamento TEXT,
        estado TEXT DEFAULT 'CONCLUIDA',
        notas TEXT,
        criado_em TEXT DEFAULT CURRENT_TIMESTAMP
    );
    CREATE TABLE IF NOT EXISTS itens_venda (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        venda_id INTEGER, vinho_id INTEGER,
        quantidade INTEGER DEFAULT 1, preco_unit REAL DEFAULT 0,
        subtotal REAL DEFAULT 0
    );
    CREATE TABLE IF NOT EXISTS caves (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nome TEXT,
        localizacao TEXT,
        capacidade INTEGER DEFAULT 0,
        temperatura_ideal REAL,
        humidade_ideal REAL,
        notas TEXT,
        ativo INTEGER DEFAULT 1,
        criado_em TEXT DEFAULT CURRENT_TIMESTAMP
    );
    CREATE TABLE IF NOT EXISTS movimentos_stock (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        vinho_id INTEGER,
        tipo TEXT,
        quantidade INTEGER,
        motivo TEXT,
        funcionario_id INTEGER,
        criado_em TEXT DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (vinho_id) REFERENCES vinhos(id)
    );
    CREATE TABLE IF NOT EXISTS clientes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        pessoa_id INTEGER,
        nif TEXT,
        data_nascimento TEXT,
        preferencias TEXT,
        ativo INTEGER DEFAULT 1,
        criado_em TEXT DEFAULT CURRENT_TIMESTAMP
    );
    """)
    if conn.execute("SELECT COUNT(*) FROM utilizadores").fetchone()[0] == 0:
        _seed_sqlite(conn)

def _seed_sqlite(conn):
    """Preencher dados iniciais SQLite"""
    conn.executemany("INSERT INTO pessoas (nome,email,telefone) VALUES (?,?,?)", [
        ("António Ferreira","gerente@vinhadouro.pt","912000001"),
        ("Sofia Martins","loja@vinhadouro.pt","912000002"),
        ("João Rodrigues","stock@vinhadouro.pt","912000003"),
        ("Admin Sistema","admin@vinhadouro.pt","912000004"),
    ])
    conn.executemany("INSERT INTO utilizadores (pessoa_id,username,password_hash,role) VALUES (?,?,?,?)", [
        (1, "gerente", generate_password_hash("1234"), "GERENTE"),
        (2, "loja",    generate_password_hash("1234"), "FUNCIONARIO"),
        (3, "stock",   generate_password_hash("1234"), "ARMAZENISTA"),
        (4, "admin",   generate_password_hash("1234"), "ADMIN"),
    ])
    vinhos = [
        # Tintos
        ("Barca Velha","Tinto","Douro","Casa Ferreirinha",2011,280.00,3,"O ícone do Douro.",None),
        ("Quinta do Crasto Reserva","Tinto","Douro","Quinta do Crasto",2019,28.50,28,"Frutos vermelhos, taninos sedosos.",None),
        ("Niepoort Redoma Tinto","Tinto","Douro","Niepoort",2020,27.00,15,"Elegante, terroir duriense.",None),
        ("Mouchão Tonel 3-4","Tinto","Alentejo","Herdade do Mouchão",2015,62.00,5,"Castelão histórico.",None),
        ("Esporão Reserva","Tinto","Alentejo","Herdade do Esporão",2019,18.90,40,"Aragonez e Trincadeira.",None),
        ("Quinta da Pellada Primus","Tinto","Dão","Quinta da Pellada",2018,32.00,11,"Touriga Nacional.",None),
        ("Luís Pato Vinhas Velhas","Tinto","Bairrada","Luís Pato",2018,22.00,16,"Baga clássica.",None),
        # Brancos
        ("Anselmo Mendes Alvarinho","Branco","Vinho Verde","Anselmo Mendes",2022,16.50,30,"Frescura mineral.",None),
        ("Niepoort Redoma Branco","Branco","Douro","Niepoort",2021,24.50,18,"Branco de altitude.",None),
        ("Esporão Branco Reserva","Branco","Alentejo","Herdade do Esporão",2022,14.80,35,"Tropical com mineral.",None),
        ("Casa de Santar Reserva","Branco","Dão","Casa de Santar",2021,12.90,28,"Encruzado equilibrado.",None),
        # Rosé
        ("Quinta dos Murças Rosé","Rosé","Douro","Quinta dos Murças",2023,12.50,32,"Rosé do Douro.",None),
        ("Mateus Rosé","Rosé","Bairrada","Sogrape",2023,5.90,80,"Ícone português.",None),
        # Espumante
        ("Raposeira Super Reserva","Espumante","Bairrada","Raposeira",2018,12.40,40,"Bagas finas.",None),
        ("Murganheira Bruto","Espumante","Távora-Varosa","Murganheira",2019,15.50,22,"Método clássico.",None),
        # Porto
        ("Graham's 20 Years Old Tawny","Porto","Douro","Graham's",None,68.00,12,"Tawny envelhecido.",None),
        ("Taylor's Vintage 2017","Porto","Douro","Taylor's",2017,135.00,6,"Vintage histórico.",None),
        ("Niepoort LBV","Porto","Douro","Niepoort",2018,22.50,24,"Late Bottled Vintage.",None),
        ("Ferreira Dona Antónia Reserva","Porto","Douro","Sogevinus",None,18.90,30,"Tawny Reserva.",None),
        ("Quinta do Noval LBV","Porto","Douro","Quinta do Noval",2017,28.00,16,"Vinhedo singular.",None),
        # Madeira
        ("Blandy's 10 Years Malmsey","Madeira","Madeira","Blandy's",None,45.00,10,"Doce, complexo.",None),
        ("Henriques & Henriques 15 Anos Sercial","Madeira","Madeira","Henriques & Henriques",None,55.00,8,"Seco, frutos secos.",None),
        ("Barbeito Verdelho 5 Anos","Madeira","Madeira","Barbeito",None,28.00,14,"Meio-seco. Caramelo salgado.",None),
        ("D'Oliveiras Boal 1989","Madeira","Madeira","D'Oliveiras",1989,180.00,3,"Colheita raríssima.",None),
    ]
    conn.executemany("INSERT INTO vinhos (nome,tipo,regiao,produtor,ano_colheita,preco,quantidade,descricao,imagem_url) VALUES (?,?,?,?,?,?,?,?,?)", vinhos)
    pessoas = [
        ("António Ferreira","antonio@vinhadouro.pt","912345678"),
        ("Sofia Martins","sofia@vinhadouro.pt","913456789"),
        ("João Rodrigues","joao@vinhadouro.pt","914567890"),
        ("Maria Costa","maria@vinhadouro.pt","915678901"),
    ]
    ids = []
    for p in pessoas:
        c = conn.execute("INSERT INTO pessoas (nome,email,telefone) VALUES (?,?,?)", p)
        ids.append(c.lastrowid)
    conn.executemany("INSERT INTO funcionarios (pessoa_id,cargo,salario,ativo,nivel) VALUES (?,?,?,?,?)", [
        (ids[0],"Gerente",2800.00,1,"GERENTE"),
        (ids[1],"Operador POS",1200.00,1,"FUNCIONARIO"),
        (ids[2],"Armazenista",1150.00,1,"ARMAZENISTA"),
        (ids[3],"Operador POS",1200.00,1,"FUNCIONARIO"),
    ])
    metodos = ["CARTAO","MB_WAY","DINHEIRO","CARTAO"]
    for i in range(8):
        dias = random.randint(0,30)
        dt = (datetime.datetime.now()-datetime.timedelta(days=dias)).strftime("%Y-%m-%d %H:%M:%S")
        total = round(random.uniform(15,200),2)
        c2 = conn.execute("INSERT INTO vendas (funcionario_id,criado_em,metodo_pagamento,estado,total) VALUES (?,?,?,?,?)",
                           (1,dt,metodos[i%4],"CONCLUIDA",total))
        conn.execute("INSERT INTO itens_venda (venda_id,vinho_id,quantidade,preco_unit,subtotal) VALUES (?,?,?,?,?)",
                     (c2.lastrowid,(i%24)+1,1,total,total))
    conn.commit()

# ── Static Files ──────────────────────────────────────────
@app.route('/')
def index():
    """Servir página principal (frontend/index.html)."""
    return send_from_directory(FRONTEND_DIR, 'index.html')

@app.route('/<path:filename>')
def serve_static(filename):
    """Servir ficheiros estáticos (frontend/...). Ignora rotas da API."""
    if filename.startswith('api/'):
        return jsonify({'error': 'Endpoint não encontrado'}), 404
    return send_from_directory(FRONTEND_DIR, filename)

# ── Health ────────────────────────────────────────────────
@app.route('/api/health')
def health():
    """Verificar saúde da API."""
    try:
        conn, db_type = get_db()
        conn.close()
        return jsonify({'status': 'ok', 'db': db_type, 'version': '3.1'})
    except Exception as e:
        log.error("Health check falhou: %s", e)
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.errorhandler(404)
def not_found(e):
    """Handler para rotas não encontradas."""
    return jsonify({'error': 'Endpoint não encontrado'}), 404

@app.errorhandler(405)
def method_not_allowed(e):
    """Handler para métodos HTTP não permitidos."""
    return jsonify({'error': 'Método não permitido'}), 405

# ── AUTH ──────────────────────────────────────────────────
def _verify_password(stored_hash, password):
    """Compatível com hash werkzeug (formato: pbkdf2:sha256:...) e com legacy plaintext."""
    if not stored_hash:
        return False
    if stored_hash.startswith('pbkdf2:') or stored_hash.startswith('scrypt:') or stored_hash.startswith('argon2:'):
        try:
            return check_password_hash(stored_hash, password)
        except Exception:
            return False
    # Legacy: plaintext (para BD antiga ainda não migrada)
    return stored_hash == password

def _maybe_upgrade_password(conn, db_type, user_id, plain):
    """Quando um utilizador faz login com sucesso e a password ainda está em plaintext,
    aproveitamos para fazer o upgrade para hash."""
    try:
        new_hash = generate_password_hash(plain)
        db_exec(conn, db_type, "UPDATE utilizadores SET password_hash=? WHERE id=?", (new_hash, user_id))
        if hasattr(conn, 'commit'):
            conn.commit()
    except Exception as e:
        log.warning("Falha ao migrar password do utilizador #%s: %s", user_id, e)

@app.route('/api/login', methods=['POST'])
def login():
    """Autenticar utilizador (hash com werkzeug, com fallback para legacy plaintext)"""
    try:
        d = request.get_json() or {}
        username = (d.get('username') or '').strip().lower()
        password = d.get('password') or ''
        if not username or not password:
            return jsonify({'error': 'Credenciais em falta'}), 400
        if len(username) > 100:
            return jsonify({'error': 'Username demasiado longo'}), 400
        conn, db_type = get_db()
        u = db_fetchone(conn, db_type,
            "SELECT u.id, u.username, u.password_hash, u.role, u.ativo, p.nome "
            "FROM utilizadores u LEFT JOIN pessoas p ON u.pessoa_id = p.id "
            "WHERE LOWER(u.username)=?",
            (username,))
        if not u or not _verify_password(u['password_hash'], password):
            conn.close()
            return jsonify({'error': 'Utilizador ou senha incorretos'}), 401
        if not u.get('ativo', 1):
            conn.close()
            return jsonify({'error': 'Conta desativada. Contacte o gerente.'}), 403
        # Auto-upgrade da password se ainda estiver em plaintext
        if u['password_hash'] and not (u['password_hash'].startswith('pbkdf2:') or
                                       u['password_hash'].startswith('scrypt:') or
                                       u['password_hash'].startswith('argon2:')):
            _maybe_upgrade_password(conn, db_type, u['id'], password)
        conn.close()
        role_redirects = {
            'GERENTE': 'gerente.html',
            'FUNCIONARIO': 'loja.html',
            'ARMAZENISTA': 'stock.html',
            'ADMIN': 'gerente.html'
        }
        token, expires = _new_token(u)
        log.info("Login: %s (%s)", username, u['role'])
        return jsonify({
            'id': u['id'],
            'username': u['username'],
            'nome': u.get('nome') or u['username'],
            'role': u['role'],
            'redirect': role_redirects.get(u['role'], 'loja.html'),
            'token': token,
            'expiresAt': expires.isoformat() + 'Z',
        })
    except Exception as e:
        log.error("Erro no login: %s", e)
        return jsonify({'error': str(e)}), 500

@app.route('/api/logout', methods=['POST'])
def logout():
    """Terminar sessão (revoga o token, se existir)"""
    token = _get_request_token()
    if token:
        _revoke_token(token)
    return jsonify({'ok': True})

@app.route('/api/me', methods=['GET'])
@requires_auth
def me():
    """Devolve informação do utilizador autenticado (para renovação de sessão)."""
    s = g.user
    return jsonify({
        'id': s['user_id'],
        'username': s['username'],
        'nome': s['nome'],
        'role': s['role'],
        'expiresAt': s['expires_at'].isoformat() + 'Z',
    })

# ── VINHOS ────────────────────────────────────────────────
@app.route('/api/vinhos', methods=['GET'])
def listar_vinhos():
    """Listar todos os vinhos"""
    try:
        conn, db_type = get_db()
        rows = db_fetchall(conn, db_type, "SELECT * FROM vinhos ORDER BY nome")
        conn.close()
        result = []
        for r in rows:
            result.append({
                'id':          r['id'],
                'nome':        r['nome'],
                'tipo':        r['tipo'],
                'regiao':      r['regiao'],
                'produtor':    r['produtor'],
                'anoColheita': r['ano_colheita'],
                'preco':       float(r['preco'] or 0),
                'quantidade':  r['quantidade'] or 0,
                'descricao':   r.get('descricao') or '',
                'imagemUrl':   r.get('imagem_url') or '',
            })
        return jsonify(result)
    except Exception as e:
        return jsonify({'error':str(e)}), 500

@app.route('/api/vinhos/<int:vid>', methods=['GET'])
def obter_vinho(vid):
    """Obter detalhes de um vinho por id."""
    try:
        conn, db_type = get_db()
        r = db_fetchone(conn, db_type, "SELECT * FROM vinhos WHERE id=?", (vid,))
        conn.close()
        if not r:
            return jsonify({'error': 'Vinho não encontrado'}), 404
        return jsonify({
            'id':          r['id'],
            'nome':        r['nome'],
            'tipo':        r['tipo'],
            'regiao':      r['regiao'],
            'produtor':    r['produtor'],
            'anoColheita': r.get('ano_colheita'),
            'preco':       float(r['preco'] or 0),
            'quantidade':  r['quantidade'] or 0,
            'descricao':   r.get('descricao') or '',
            'imagemUrl':   r.get('imagem_url') or '',
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/vinhos', methods=['POST'])
@requires_role('GERENTE', 'ADMIN', 'ARMAZENISTA')
def criar_vinho():
    """Criar novo vinho"""
    try:
        d = request.get_json()
        if not d.get('nome'):
            return jsonify({'error':'Nome é obrigatório'}), 400
        conn, db_type = get_db()
        if db_type == 'mysql':
            with conn.cursor() as c:
                c.execute("""INSERT INTO vinhos (nome,tipo,regiao,produtor,ano_colheita,preco,quantidade,descricao)
                             VALUES (%s,%s,%s,%s,%s,%s,%s,%s)""",
                          (d.get('nome'),d.get('tipo'),d.get('regiao'),d.get('produtor'),
                           d.get('anoColheita'),d.get('preco',0),d.get('quantidade',0),d.get('descricao','')))
                new_id = c.lastrowid
            conn.commit()
        else:
            c = conn.execute("INSERT INTO vinhos (nome,tipo,regiao,produtor,ano_colheita,preco,quantidade,descricao) VALUES (?,?,?,?,?,?,?,?)",
                             (d.get('nome'),d.get('tipo'),d.get('regiao'),d.get('produtor'),
                              d.get('anoColheita'),d.get('preco',0),d.get('quantidade',0),d.get('descricao','')))
            new_id = c.lastrowid
            conn.commit()
        conn.close()
        return jsonify({'id':new_id,**d}), 201
    except Exception as e:
        return jsonify({'error':str(e)}), 500

@app.route('/api/vinhos/<int:vid>', methods=['PUT'])
@requires_role('GERENTE', 'ADMIN', 'ARMAZENISTA')
def atualizar_vinho(vid):
    """Atualizar vinho existente"""
    try:
        d = request.get_json()
        conn, db_type = get_db()
        db_exec(conn, db_type, """UPDATE vinhos SET nome=?,tipo=?,regiao=?,produtor=?,
                                   ano_colheita=?,preco=?,quantidade=?,descricao=? WHERE id=?""",
                (d.get('nome'),d.get('tipo'),d.get('regiao'),d.get('produtor'),
                 d.get('anoColheita'),d.get('preco',0),d.get('quantidade',0),d.get('descricao',''),vid))
        conn.commit()
        conn.close()
        return jsonify({'id':vid,**d})
    except Exception as e:
        return jsonify({'error':str(e)}), 500

@app.route('/api/vinhos/<int:vid>', methods=['DELETE'])
@requires_role('GERENTE', 'ADMIN')
def apagar_vinho(vid):
    """Apagar vinho"""
    try:
        conn, db_type = get_db()
        db_exec(conn, db_type, "DELETE FROM vinhos WHERE id=?", (vid,))
        conn.commit()
        conn.close()
        return jsonify({'ok':True})
    except Exception as e:
        return jsonify({'error':str(e)}), 500

@app.route('/api/vinhos/<int:vid>/stock', methods=['PUT'])
@requires_role('GERENTE', 'ADMIN', 'ARMAZENISTA')
def atualizar_stock(vid):
    """Atualizar quantidade em stock"""
    try:
        d = request.get_json()
        qty = d.get('quantidade', 0)
        if qty < 0:
            return jsonify({'error':'Quantidade não pode ser negativa'}), 400
        conn, db_type = get_db()
        db_exec(conn, db_type, "UPDATE vinhos SET quantidade=? WHERE id=?", (qty, vid))
        conn.commit()
        conn.close()
        return jsonify({'id':vid,'quantidade':qty})
    except Exception as e:
        return jsonify({'error':str(e)}), 500

@app.route('/api/vinhos/stock-critico')
def stock_critico():
    """Listar vinhos com stock crítico"""
    try:
        conn, db_type = get_db()
        rows = db_fetchall(conn, db_type, "SELECT * FROM vinhos WHERE quantidade < 10 ORDER BY quantidade")
        conn.close()
        return jsonify([{'id':r['id'],'nome':r['nome'],'quantidade':r['quantidade'],'tipo':r['tipo']} for r in rows])
    except Exception as e:
        return jsonify({'error':str(e)}), 500

# ── FUNCIONÁRIOS ──────────────────────────────────────────
@app.route('/api/funcionarios', methods=['GET'])
def listar_funcionarios():
    """Listar todos os funcionários"""
    try:
        conn, db_type = get_db()
        rows = db_fetchall(conn, db_type, """
            SELECT f.*,p.nome,p.email,p.telefone FROM funcionarios f
            LEFT JOIN pessoas p ON f.pessoa_id=p.id ORDER BY p.nome
        """)
        conn.close()
        result = []
        for r in rows:
            result.append({
                'id':          r['id'],
                'nome':        r['nome'],
                'email':       r.get('email',''),
                'telefone':    r.get('telefone',''),
                'cargo':       r['cargo'],
                'salario':     float(r['salario'] or 0),
                'dataAdmissao':str(r.get('criado_em')) if r.get('criado_em') else '',
                'ativo':       r['ativo'],
                'nivelAcesso': r['nivel'],
            })
        return jsonify(result)
    except Exception as e:
        return jsonify({'error':str(e)}), 500

@app.route('/api/funcionarios', methods=['POST'])
@requires_role('GERENTE', 'ADMIN')
def criar_funcionario():
    """Criar novo funcionário"""
    try:
        d = request.get_json()
        if not d.get('nome'):
            return jsonify({'error':'Nome é obrigatório'}), 400
        conn, db_type = get_db()
        if db_type == 'mysql':
            with conn.cursor() as c:
                c.execute("INSERT INTO pessoas (nome,email) VALUES (%s,%s)", (d.get('nome'),d.get('email','')))
                pid = c.lastrowid
                c.execute("INSERT INTO funcionarios (pessoa_id,cargo,salario,ativo,nivel) VALUES (%s,%s,%s,%s,%s)",
                          (pid,d.get('cargo'),d.get('salario',0),1,d.get('nivelAcesso','FUNCIONARIO')))
                new_id = c.lastrowid
            conn.commit()
        else:
            c = conn.execute("INSERT INTO pessoas (nome,email) VALUES (?,?)", (d.get('nome'),d.get('email','')))
            pid = c.lastrowid
            c2 = conn.execute("INSERT INTO funcionarios (pessoa_id,cargo,salario,ativo,nivel) VALUES (?,?,?,?,?)",
                              (pid,d.get('cargo'),d.get('salario',0),1,d.get('nivelAcesso','FUNCIONARIO')))
            new_id = c2.lastrowid
            conn.commit()
        conn.close()
        return jsonify({'id':new_id,**d}), 201
    except Exception as e:
        return jsonify({'error':str(e)}), 500

@app.route('/api/funcionarios/<int:fid>', methods=['PUT'])
@requires_role('GERENTE', 'ADMIN')
def atualizar_funcionario(fid):
    """Atualizar funcionário"""
    try:
        d = request.get_json()
        conn, db_type = get_db()
        row = db_fetchone(conn, db_type, "SELECT pessoa_id FROM funcionarios WHERE id=?", (fid,))
        if row:
            pid = row['pessoa_id']
            db_exec(conn, db_type, "UPDATE pessoas SET nome=?,email=? WHERE id=?", (d.get('nome'),d.get('email',''),pid))
            db_exec(conn, db_type, "UPDATE funcionarios SET cargo=?,salario=?,ativo=?,nivel=? WHERE id=?",
                    (d.get('cargo'),d.get('salario',0),d.get('ativo',1),d.get('nivelAcesso','FUNCIONARIO'),fid))
        conn.commit()
        conn.close()
        return jsonify({'id':fid,**d})
    except Exception as e:
        return jsonify({'error':str(e)}), 500

# ── VENDAS ────────────────────────────────────────────────
@app.route('/api/vendas', methods=['GET'])
def listar_vendas():
    """Listar todas as vendas — devolve cliente E funcionário separados (ambos via JOINs distintos)."""
    try:
        conn, db_type = get_db()
        # Cliente vem de v.cliente_id -> clientes.pessoa_id -> pessoas.nome (consumidor final)
        # Funcionário vem de v.funcionario_id -> funcionarios.pessoa_id -> pessoas.nome (operador POS)
        sep = "SEPARATOR ', '" if db_type == 'mysql' else ", ', '"
        rows = db_fetchall(conn, db_type, f"""
            SELECT v.id, v.codigo, v.criado_em, v.metodo_pagamento, v.estado, v.total,
                   pc.nome  AS cliente_nome,
                   c.nif    AS cliente_nif,
                   pf.nome  AS funcionario_nome,
                   GROUP_CONCAT(vi.nome {sep}) AS produtos_nomes
            FROM vendas v
            LEFT JOIN clientes c       ON v.cliente_id = c.id
            LEFT JOIN pessoas pc       ON c.pessoa_id = pc.id
            LEFT JOIN funcionarios f   ON v.funcionario_id = f.id
            LEFT JOIN pessoas pf       ON f.pessoa_id = pf.id
            LEFT JOIN itens_venda iv   ON iv.venda_id = v.id
            LEFT JOIN vinhos vi        ON iv.vinho_id = vi.id
            GROUP BY v.id ORDER BY v.criado_em DESC LIMIT 100
        """)
        conn.close()
        result = []
        for r in rows:
            result.append({
                'id':               r['id'],
                'codigo':           r.get('codigo') or f"VD-{r['id']:04d}",
                'cliente':          r.get('cliente_nome') or 'Consumidor Final',
                'clienteNome':      r.get('cliente_nome') or 'Consumidor Final',
                'clienteNif':       r.get('cliente_nif') or '',
                'funcionarioNome':  r.get('funcionario_nome') or 'Sistema',
                'produto':          r.get('produtos_nomes') or '—',
                'metodoPagamento':  r['metodo_pagamento'],
                'total':            float(r['total'] or 0),
                'status':           r['estado'],
                'estado':           r['estado'],
                'dataVenda':        str(r['criado_em']) if r.get('criado_em') else '',
            })
        return jsonify(result)
    except Exception as e:
        log.error("Erro a listar vendas: %s", e)
        return jsonify({'error':str(e)}), 500

def _resolve_or_create_cliente(conn, db_type, dados):
    """Localiza ou cria cliente a partir de dados básicos. Devolve cliente_id ou None.
    `dados` é um dict opcional com {nome, email, telefone, nif, morada}.
    Se não houver nome nem nif, devolve None (consumidor final)."""
    if not dados:
        return None
    nome = (dados.get('nome') or '').strip()
    nif = (dados.get('nif') or '').strip()
    email = (dados.get('email') or '').strip()
    telefone = (dados.get('telefone') or '').strip()
    if not nome and not nif:
        return None  # consumidor final
    # Procurar cliente existente pelo NIF
    if nif:
        row = db_fetchone(conn, db_type, "SELECT id FROM clientes WHERE nif=? LIMIT 1", (nif,))
        if row:
            return row['id']
    # Criar pessoa + cliente
    if db_type == 'mysql':
        with conn.cursor() as c:
            c.execute("INSERT INTO pessoas (nome,email,telefone) VALUES (%s,%s,%s)",
                      (nome or 'Cliente', email or None, telefone or None))
            pid = c.lastrowid
            c.execute("INSERT INTO clientes (pessoa_id,nif,ativo) VALUES (%s,%s,1)",
                      (pid, nif or None))
            cid = c.lastrowid
    else:
        cur = conn.execute("INSERT INTO pessoas (nome,email,telefone) VALUES (?,?,?)",
                           (nome or 'Cliente', email or None, telefone or None))
        pid = cur.lastrowid
        cur2 = conn.execute("INSERT INTO clientes (pessoa_id,nif,ativo) VALUES (?,?,1)",
                            (pid, nif or None))
        cid = cur2.lastrowid
    return cid

@app.route('/api/vendas', methods=['POST'])
@requires_auth
def criar_venda():
    """Criar nova venda — usa o utilizador autenticado como funcionário e
    aceita opcionalmente dados de cliente para registar consumidor final identificado."""
    try:
        d = request.get_json() or {}
        itens    = d.get('itens', [])
        metodo_raw = d.get('metodoPagamento', 'DINHEIRO')
        metodo_key = (metodo_raw if isinstance(metodo_raw, str) else str(metodo_raw)).strip().lower()
        metodo_map = {
            'cartão': 'CARTAO', 'cartao': 'CARTAO', 'CARTAO': 'CARTAO',
            'numerário': 'DINHEIRO', 'numerario': 'DINHEIRO', 'dinheiro': 'DINHEIRO', 'DINHEIRO': 'DINHEIRO',
            'mb way': 'MB_WAY', 'mb_way': 'MB_WAY', 'mbway': 'MB_WAY', 'MB_WAY': 'MB_WAY',
            'transferência': 'TRANSFERENCIA', 'transferencia': 'TRANSFERENCIA', 'TRANSFERENCIA': 'TRANSFERENCIA',
            'multibanco': 'TRANSFERENCIA',
        }
        metodo = metodo_map.get(metodo_key, 'DINHEIRO')

        # Funcionário: prefere o utilizador autenticado (g.user). Faz match user_id->funcionario_id via pessoa_id.
        func_id = d.get('funcionarioId')
        if not func_id:
            try:
                user_id = g.user['user_id']
            except Exception:
                user_id = None
            if user_id:
                conn_tmp, dbt_tmp = get_db()
                row = db_fetchone(conn_tmp, dbt_tmp,
                    "SELECT f.id FROM funcionarios f "
                    "JOIN utilizadores u ON u.pessoa_id = f.pessoa_id "
                    "WHERE u.id=? LIMIT 1", (user_id,))
                conn_tmp.close()
                if row:
                    func_id = row['id']
        if not func_id:
            func_id = 1  # fallback (utilizador admin/sistema)

        if d.get('vinhoId') and not itens:
            itens = [{'vinhoId': d['vinhoId'], 'quantidade': d.get('quantidade', 1)}]

        if not itens:
            return jsonify({'error': 'Venda deve conter pelo menos um item'}), 400

        conn, db_type = get_db()

        # Cliente: opcional. Se vier `cliente` (objecto) ou `clienteNome`+`clienteNif` em flat.
        cliente_dados = d.get('cliente')
        if not cliente_dados and (d.get('clienteNome') or d.get('clienteNif')):
            cliente_dados = {
                'nome':     d.get('clienteNome'),
                'nif':      d.get('clienteNif'),
                'email':    d.get('clienteEmail'),
                'telefone': d.get('clienteTelefone'),
            }
        cliente_id = _resolve_or_create_cliente(conn, db_type, cliente_dados)

        total = 0.0
        item_data = []
        for item in itens:
            vid = item.get('vinhoId') or item.get('vinho_id')
            qty = item.get('quantidade', 1)
            if not vid or qty < 1:
                conn.close()
                return jsonify({'error': 'Item inválido: vinhoId e quantidade são obrigatórios'}), 400
            row = db_fetchone(conn, db_type, "SELECT preco, quantidade FROM vinhos WHERE id=?", (vid,))
            if not row:
                conn.close()
                return jsonify({'error': f'Vinho #{vid} não encontrado'}), 404
            stock_atual = row['quantidade'] or 0
            if qty > stock_atual:
                conn.close()
                return jsonify({'error': f'Stock insuficiente para vinho #{vid} (disponível: {stock_atual})'}), 400
            preco = float(row['preco'] or 0)
            total += preco * qty
            item_data.append((vid, qty, preco))
            new_stock = stock_atual - qty
            db_exec(conn, db_type, "UPDATE vinhos SET quantidade=? WHERE id=?", (new_stock, vid))

        notas = (d.get('notas') or '').strip() or None
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if db_type == 'mysql':
            with conn.cursor() as c:
                c.execute(
                    "INSERT INTO vendas (cliente_id,funcionario_id,criado_em,metodo_pagamento,estado,total,notas) "
                    "VALUES (%s,%s,%s,%s,%s,%s,%s)",
                    (cliente_id, func_id, now, metodo, 'CONCLUIDA', round(total, 2), notas)
                )
                venda_id = c.lastrowid
                # Atribuir código legível (VD-2026-XXXX)
                codigo = f"VD-{datetime.datetime.now().year}-{venda_id:04d}"
                c.execute("UPDATE vendas SET codigo=%s WHERE id=%s", (codigo, venda_id))
                for (vid, qty, preco) in item_data:
                    subtotal = round(preco * qty, 2)
                    c.execute(
                        "INSERT INTO itens_venda (venda_id,vinho_id,quantidade,preco_unit,subtotal) VALUES (%s,%s,%s,%s,%s)",
                        (venda_id, vid, qty, preco, subtotal)
                    )
            conn.commit()
        else:
            c = conn.execute(
                "INSERT INTO vendas (cliente_id,funcionario_id,criado_em,metodo_pagamento,estado,total,notas) VALUES (?,?,?,?,?,?,?)",
                (cliente_id, func_id, now, metodo, 'CONCLUIDA', round(total, 2), notas)
            )
            venda_id = c.lastrowid
            codigo = f"VD-{datetime.datetime.now().year}-{venda_id:04d}"
            conn.execute("UPDATE vendas SET codigo=? WHERE id=?", (codigo, venda_id))
            for (vid, qty, preco) in item_data:
                subtotal = round(preco * qty, 2)
                conn.execute(
                    "INSERT INTO itens_venda (venda_id,vinho_id,quantidade,preco_unit,subtotal) VALUES (?,?,?,?,?)",
                    (venda_id, vid, qty, preco, subtotal)
                )
            conn.commit()
        conn.close()
        return jsonify({
            'id': venda_id,
            'codigo': codigo,
            'total': round(total, 2),
            'status': 'CONCLUIDA',
            'estado': 'CONCLUIDA',
            'metodoPagamento': metodo,
            'dataVenda': now,
            'clienteId': cliente_id,
            'funcionarioId': func_id,
        }), 201
    except Exception as e:
        log.error("Erro a criar venda: %s", e)
        return jsonify({'error': str(e)}), 500

# ── CAVES (Wine Cellars) ──────────────────────────────────
@app.route('/api/caves', methods=['GET'])
def listar_caves():
    """Listar todas as caves"""
    try:
        conn, db_type = get_db()
        rows = db_fetchall(conn, db_type, "SELECT * FROM caves WHERE ativo=1 ORDER BY nome")
        conn.close()
        result = []
        for r in rows:
            result.append({
                'id':               r['id'],
                'nome':             r['nome'],
                'localizacao':      r.get('localizacao',''),
                'capacidade':       r['capacidade'],
                'temperaturaIdeal': float(r['temperatura_ideal']) if r.get('temperatura_ideal') else None,
                'humidadeIdeal':    float(r['humidade_ideal']) if r.get('humidade_ideal') else None,
                'notas':            r.get('notas',''),
            })
        return jsonify(result)
    except Exception as e:
        return jsonify({'error':str(e)}), 500

@app.route('/api/caves', methods=['POST'])
@requires_role('GERENTE', 'ADMIN', 'ARMAZENISTA')
def criar_cave():
    """Criar nova cave"""
    try:
        d = request.get_json()
        if not d.get('nome'):
            return jsonify({'error':'Nome é obrigatório'}), 400

        conn, db_type = get_db()
        db_exec(conn, db_type,
                "INSERT INTO caves (nome,localizacao,capacidade,temperatura_ideal,humidade_ideal,notas,ativo) VALUES (?,?,?,?,?,?,?)",
                (d.get('nome'), d.get('localizacao',''), d.get('capacidade',100),
                 d.get('temperaturaIdeal'), d.get('humidadeIdeal'), d.get('notas',''), 1))
        conn.commit()
        conn.close()
        return jsonify({'ok':True}), 201
    except Exception as e:
        return jsonify({'error':str(e)}), 500

@app.route('/api/caves/<int:cave_id>', methods=['PUT'])
@requires_role('GERENTE', 'ADMIN', 'ARMAZENISTA')
def atualizar_cave(cave_id):
    """Atualizar cave"""
    try:
        d = request.get_json()
        conn, db_type = get_db()

        fields, vals = [], []
        for key, col in [('nome','nome'),('localizacao','localizacao'),('capacidade','capacidade'),
                         ('temperaturaIdeal','temperatura_ideal'),('humidadeIdeal','humidade_ideal'),('notas','notas')]:
            if key in d:
                fields.append(f"{col}=?")
                vals.append(d[key])

        if fields:
            vals.append(cave_id)
            db_exec(conn, db_type, f"UPDATE caves SET {','.join(fields)} WHERE id=?", vals)
            conn.commit()

        conn.close()
        return jsonify({'ok':True})
    except Exception as e:
        return jsonify({'error':str(e)}), 500

# ── MOVIMENTOS DE STOCK ───────────────────────────────────
@app.route('/api/movimentos-stock', methods=['GET'])
def listar_movimentos():
    """Listar movimentos de stock"""
    try:
        conn, db_type = get_db()
        rows = db_fetchall(conn, db_type, """
            SELECT m.*, v.nome as vinho_nome, f.pessoa_id, p.nome as funcionario_nome
            FROM movimentos_stock m
            LEFT JOIN vinhos v ON m.vinho_id = v.id
            LEFT JOIN funcionarios f ON m.funcionario_id = f.id
            LEFT JOIN pessoas p ON f.pessoa_id = p.id
            ORDER BY m.criado_em DESC LIMIT 100
        """)
        conn.close()
        result = []
        for r in rows:
            result.append({
                'id':               r['id'],
                'vinhoId':          r['vinho_id'],
                'vinhoNome':        r.get('vinho_nome',''),
                'tipo':             r['tipo'],
                'quantidade':       r['quantidade'],
                'motivo':           r.get('motivo',''),
                'funcionarioNome':  r.get('funcionario_nome',''),
                'criadoEm':         str(r['criado_em']) if r.get('criado_em') else '',
            })
        return jsonify(result)
    except Exception as e:
        return jsonify({'error':str(e)}), 500

@app.route('/api/movimentos-stock', methods=['POST'])
@requires_role('GERENTE', 'ADMIN', 'ARMAZENISTA')
def registar_movimento():
    """Registar novo movimento de stock"""
    try:
        d = request.get_json()
        if not d.get('vinhoId') or not d.get('tipo') or d.get('quantidade') is None:
            return jsonify({'error':'vinhoId, tipo e quantidade são obrigatórios'}), 400

        tipo_raw = d.get('tipo','').upper()
        if tipo_raw not in ['ENTRADA', 'SAIDA']:
            return jsonify({'error':'tipo deve ser ENTRADA ou SAIDA'}), 400

        conn, db_type = get_db()

        # Atualizar stock do vinho
        vinho = db_fetchone(conn, db_type, "SELECT quantidade FROM vinhos WHERE id=?", (d.get('vinhoId'),))
        if vinho:
            current_qty = vinho['quantidade'] or 0
            if tipo_raw == 'ENTRADA':
                new_qty = current_qty + d.get('quantidade',0)
            else:
                new_qty = max(0, current_qty - d.get('quantidade',0))

            db_exec(conn, db_type, "UPDATE vinhos SET quantidade=? WHERE id=?", (new_qty, d.get('vinhoId')))

        # Registar movimento
        db_exec(conn, db_type,
                "INSERT INTO movimentos_stock (vinho_id,tipo,quantidade,motivo,funcionario_id) VALUES (?,?,?,?,?)",
                (d.get('vinhoId'), tipo_raw, d.get('quantidade'), d.get('motivo',''), d.get('funcionarioId')))

        conn.commit()
        conn.close()
        return jsonify({'ok':True}), 201
    except Exception as e:
        return jsonify({'error':str(e)}), 500

# ── CLIENTES ──────────────────────────────────────────────
@app.route('/api/clientes', methods=['GET'])
def listar_clientes():
    """Listar todos os clientes"""
    try:
        conn, db_type = get_db()
        rows = db_fetchall(conn, db_type, """
            SELECT c.*, p.nome, p.email, p.telefone
            FROM clientes c
            LEFT JOIN pessoas p ON c.pessoa_id = p.id
            WHERE c.ativo = 1
            ORDER BY p.nome
        """)
        conn.close()
        result = []
        for r in rows:
            prefs_raw = r.get('preferencias')
            preferencias = {}
            if prefs_raw:
                try:
                    preferencias = json.loads(prefs_raw) if isinstance(prefs_raw, str) else (prefs_raw or {})
                    if not isinstance(preferencias, dict):
                        preferencias = {}
                except (json.JSONDecodeError, TypeError, ValueError):
                    preferencias = {}
            result.append({
                'id':               r['id'],
                'nome':             r.get('nome',''),
                'email':            r.get('email',''),
                'telefone':         r.get('telefone',''),
                'morada':           '',
                'nif':              r.get('nif',''),
                'dataNascimento':   str(r['data_nascimento']) if r.get('data_nascimento') else None,
                'preferencias':     preferencias,
            })
        return jsonify(result)
    except Exception as e:
        return jsonify({'error':str(e)}), 500

# ── RELATÓRIOS ────────────────────────────────────────────
@app.route('/api/relatorios/vendas-mensal', methods=['GET'])
def relatorio_vendas_mensal():
    """Relatório de vendas mensais"""
    try:
        conn, db_type = get_db()

        if db_type == 'mysql':
            rows = db_fetchall(conn, db_type, """
                SELECT
                    DATE_FORMAT(v.criado_em, '%Y-%m') as mes,
                    COUNT(*) as num_vendas,
                    SUM(v.total) as total_vendas,
                    AVG(v.total) as ticket_medio
                FROM vendas v
                WHERE v.estado = 'CONCLUIDA'
                GROUP BY DATE_FORMAT(v.criado_em, '%Y-%m')
                ORDER BY mes DESC
                LIMIT 12
            """)
        else:
            rows = db_fetchall(conn, db_type, """
                SELECT
                    strftime('%Y-%m', v.criado_em) as mes,
                    COUNT(*) as num_vendas,
                    SUM(v.total) as total_vendas,
                    AVG(v.total) as ticket_medio
                FROM vendas v
                WHERE v.estado = 'CONCLUIDA'
                GROUP BY strftime('%Y-%m', v.criado_em)
                ORDER BY mes DESC
                LIMIT 12
            """)

        conn.close()
        result = []
        for r in rows:
            result.append({
                'mes':          r['mes'],
                'numVendas':    r['num_vendas'],
                'totalVendas':  float(r['total_vendas'] or 0),
                'ticketMedio':  float(r['ticket_medio'] or 0),
            })
        return jsonify(result)
    except Exception as e:
        return jsonify({'error':str(e)}), 500

@app.route('/api/relatorios/top-vinhos', methods=['GET'])
def relatorio_top_vinhos():
    """Relatório dos vinhos mais vendidos"""
    try:
        conn, db_type = get_db()
        rows = db_fetchall(conn, db_type, """
            SELECT
                v.id, v.nome, v.tipo, v.regiao, v.produtor,
                COUNT(iv.id) as num_vendas,
                SUM(iv.quantidade) as quantidade_vendida,
                SUM(iv.quantidade * iv.preco_unit) as receita_total,
                v.preco
            FROM vinhos v
            LEFT JOIN itens_venda iv ON v.id = iv.vinho_id
            GROUP BY v.id, v.nome, v.tipo, v.regiao, v.produtor, v.preco
            ORDER BY receita_total DESC
            LIMIT 20
        """)
        conn.close()
        result = []
        for r in rows:
            result.append({
                'id':               r['id'],
                'nome':             r['nome'],
                'tipo':             r['tipo'],
                'regiao':           r['regiao'],
                'produtor':         r['produtor'],
                'numVendas':        r['num_vendas'] or 0,
                'quantidadeVendida': r['quantidade_vendida'] or 0,
                'receitaTotal':     float(r['receita_total'] or 0),
                'preco':            float(r['preco'] or 0),
            })
        return jsonify(result)
    except Exception as e:
        return jsonify({'error':str(e)}), 500

# ── DASHBOARD ─────────────────────────────────────────────
@app.route('/api/dashboard')
def dashboard():
    """Dashboard com estatísticas principais"""
    try:
        conn, db_type = get_db()
        receita     = (db_fetchone(conn,db_type,"SELECT COALESCE(SUM(total),0) as v FROM vendas WHERE estado='CONCLUIDA'") or {}).get('v',0)
        num_vendas  = (db_fetchone(conn,db_type,"SELECT COUNT(*) as v FROM vendas WHERE estado='CONCLUIDA'") or {}).get('v',0)
        ticket      = round(float(receita)/num_vendas,2) if num_vendas else 0
        lucro       = round(float(receita)*0.34,2)
        stock_crit  = (db_fetchone(conn,db_type,"SELECT COUNT(*) as v FROM vinhos WHERE quantidade < 10") or {}).get('v',0)
        valor_stock = (db_fetchone(conn,db_type,"SELECT COALESCE(SUM(preco*quantidade),0) as v FROM vinhos") or {}).get('v',0)
        total_vinhos= (db_fetchone(conn,db_type,"SELECT COUNT(*) as v FROM vinhos") or {}).get('v',0)
        total_func  = (db_fetchone(conn,db_type,"SELECT COUNT(*) as v FROM funcionarios WHERE ativo=1") or {}).get('v',0)

        weekly, labels = [], []
        for i in range(6,-1,-1):
            d = datetime.date.today() - datetime.timedelta(days=i)
            lbl = ["Seg","Ter","Qua","Qui","Sex","Sáb","Dom"][d.weekday()] if i>0 else "Hj"
            if db_type == 'mysql':
                total_day = (db_fetchone(conn,db_type,"SELECT COALESCE(SUM(total),0) as v FROM vendas WHERE DATE(criado_em)=%s",(d.isoformat(),)) or {}).get('v',0)
            else:
                total_day = (db_fetchone(conn,db_type,"SELECT COALESCE(SUM(total),0) as v FROM vendas WHERE date(criado_em)=?",(d.isoformat(),)) or {}).get('v',0)
            weekly.append(round(float(total_day),2))
            labels.append(lbl)

        top_vinhos = db_fetchall(conn, db_type, "SELECT nome,tipo,preco,quantidade FROM vinhos ORDER BY (preco*quantidade) DESC LIMIT 5")
        conn.close()

        return jsonify({
            'receitaTotal':   round(float(receita),2),
            'vendas':         num_vendas,
            'lucroLiquido':   lucro,
            'ticketMedio':    ticket,
            'stockCritico':   stock_crit,
            'valorStock':     round(float(valor_stock),2),
            'totalVinhos':    total_vinhos,
            'totalFuncionarios': total_func,
            'vendasSemanais': weekly,
            'vendasLabels':   labels,
            'topVinhos':      [{'nome':r['nome'],'tipo':r['tipo'],'valor':round(float(r['preco'] or 0)*int(r['quantidade'] or 0),2)} for r in top_vinhos],
        })
    except Exception as e:
        return jsonify({'error':str(e)}), 500

# ── DEVOLUÇÕES (Returns) ─────────────────────────────────
@app.route('/api/devolucoes', methods=['POST'])
@requires_auth
def criar_devolucao():
    """Processar devolução de uma venda"""
    try:
        d = request.get_json()
        venda_id = d.get('vendaId')
        motivo = d.get('motivo', '')
        itens = d.get('itens', [])  # [{vinhoId, quantidade}]

        if not venda_id:
            return jsonify({'error': 'vendaId é obrigatório'}), 400

        conn, db_type = get_db()

        # Verify venda exists
        venda = db_fetchone(conn, db_type, "SELECT * FROM vendas WHERE id=?", (venda_id,))
        if not venda:
            conn.close()
            return jsonify({'error': 'Venda não encontrada'}), 404

        total_devolvido = 0.0

        if itens:
            # Partial return - only specified items
            for item in itens:
                vid = item.get('vinhoId')
                qty = item.get('quantidade', 1)
                # Get price from itens_venda
                iv_row = db_fetchone(conn, db_type,
                    "SELECT preco_unit, quantidade FROM itens_venda WHERE venda_id=? AND vinho_id=?",
                    (venda_id, vid))
                if iv_row:
                    preco = float(iv_row['preco_unit'] or 0)
                    total_devolvido += preco * qty
                    # Restore stock
                    db_exec(conn, db_type,
                        "UPDATE vinhos SET quantidade = quantidade + ? WHERE id=?", (qty, vid))
        else:
            # Full return - return all items
            all_items = db_fetchall(conn, db_type,
                "SELECT vinho_id, quantidade, preco_unit FROM itens_venda WHERE venda_id=?",
                (venda_id,))
            for item in all_items:
                vid = item['vinho_id']
                qty = item['quantidade']
                preco = float(item['preco_unit'] or 0)
                total_devolvido += preco * qty
                # Restore stock
                db_exec(conn, db_type,
                    "UPDATE vinhos SET quantidade = quantidade + ? WHERE id=?", (qty, vid))

        # Update venda status
        db_exec(conn, db_type, "UPDATE vendas SET estado='CANCELADA' WHERE id=?", (venda_id,))
        conn.commit()
        conn.close()

        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        return jsonify({
            'ok': True,
            'vendaId': venda_id,
            'totalDevolvido': round(total_devolvido, 2),
            'motivo': motivo,
            'data': now
        }), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/vendas/<int:vid>', methods=['GET'])
def obter_venda(vid):
    """Obter detalhes de uma venda por id (para fatura/recibo)."""
    try:
        conn, db_type = get_db()
        r = db_fetchone(conn, db_type, """
            SELECT v.id, v.codigo, v.criado_em, v.metodo_pagamento, v.estado, v.total, v.notas,
                   pc.nome  as cliente_nome,
                   pc.email as cliente_email,
                   pc.telefone as cliente_telefone,
                   c.nif    as cliente_nif,
                   pf.nome  as funcionario_nome
            FROM vendas v
            LEFT JOIN clientes c       ON v.cliente_id = c.id
            LEFT JOIN pessoas pc       ON c.pessoa_id = pc.id
            LEFT JOIN funcionarios f   ON v.funcionario_id = f.id
            LEFT JOIN pessoas pf       ON f.pessoa_id = pf.id
            WHERE v.id=?
        """, (vid,))
        conn.close()
        if not r:
            return jsonify({'error': 'Venda não encontrada'}), 404
        return jsonify({
            'id':               r['id'],
            'codigo':           r.get('codigo') or f"VD-{r['id']:04d}",
            'clienteNome':      r.get('cliente_nome') or 'Consumidor Final',
            'clienteNif':       r.get('cliente_nif') or '',
            'clienteEmail':     r.get('cliente_email') or '',
            'clienteTelefone':  r.get('cliente_telefone') or '',
            'clienteMorada':    '',
            'metodoPagamento':  r['metodo_pagamento'],
            'total':            float(r['total'] or 0),
            'status':           r['estado'],
            'estado':           r['estado'],
            'dataVenda':        str(r['criado_em']) if r.get('criado_em') else '',
            'funcionarioNome':  r.get('funcionario_nome') or 'Sistema',
            'notas':            r.get('notas') or '',
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/vendas/<int:vid>/itens', methods=['GET'])
def itens_venda(vid):
    """Obter itens de uma venda específica (com detalhes do vinho para fatura)."""
    try:
        conn, db_type = get_db()
        rows = db_fetchall(conn, db_type, """
            SELECT iv.vinho_id, iv.quantidade, iv.preco_unit,
                   v.nome as vinho_nome, v.tipo, v.regiao, v.ano_colheita
            FROM itens_venda iv
            LEFT JOIN vinhos v ON iv.vinho_id = v.id
            WHERE iv.venda_id = ?
        """, (vid,))
        conn.close()
        return jsonify([{
            'vinhoId':       r['vinho_id'],
            'nome':          r.get('vinho_nome', ''),
            'tipo':          r.get('tipo', ''),
            'regiao':        r.get('regiao', ''),
            'anoColheita':   r.get('ano_colheita'),
            'quantidade':    r['quantidade'],
            'precoUnitario': float(r['preco_unit'] or 0),
        } for r in rows])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── EXPORTS XLSX (premium, formatados) ────────────────────
from io import BytesIO
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
    from openpyxl.utils import get_column_letter
    from flask import send_file
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    log.warning("openpyxl não instalado — endpoints /api/export/*.xlsx desativados")

# Paleta the 100's
_BRAND_GOLD = 'C9A96E'
_BRAND_GOLD_LIGHT = 'F2E8D0'
_BRAND_INK = '15110D'
_BRAND_CREAM = 'F5F0E8'
_BRAND_MUTED = '8B7C68'

def _fmt_header_row(ws, row_idx, columns, fill_hex=_BRAND_INK, font_color=_BRAND_GOLD):
    fill = PatternFill('solid', fgColor=fill_hex)
    font = Font(name='Inter', bold=True, color=font_color, size=10)
    align = Alignment(horizontal='left', vertical='center', wrap_text=False)
    for col_idx, _ in enumerate(columns, 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
    ws.row_dimensions[row_idx].height = 28

def _add_brand_header(ws, title, subtitle=None):
    """Bloco de cabeçalho com nome da marca, título do relatório, data."""
    # Linha 1: marca
    ws.cell(row=1, column=1, value="the 100's").font = Font(name='Cormorant Garamond', bold=True, size=22, color=_BRAND_GOLD)
    ws.cell(row=1, column=2, value="Bottled Memories").font = Font(name='Cormorant Garamond', italic=True, size=12, color=_BRAND_MUTED)
    ws.row_dimensions[1].height = 34
    # Linha 2: título
    ws.cell(row=2, column=1, value=title).font = Font(name='Inter', bold=True, size=14, color=_BRAND_INK)
    # Linha 3: subtítulo + data
    info = subtitle or ''
    if info: info += '  ·  '
    info += f"Gerado em {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws.cell(row=3, column=1, value=info).font = Font(name='Inter', size=10, color=_BRAND_MUTED)
    # Espaço
    ws.row_dimensions[4].height = 8

@app.route('/api/export/vendas.xlsx', methods=['GET'])
def export_vendas_xlsx():
    """Exporta histórico de vendas como ficheiro Excel premium (the 100's branded)."""
    if not OPENPYXL_AVAILABLE:
        return jsonify({'error': 'openpyxl não está instalado no servidor'}), 503
    try:
        conn, db_type = get_db()
        sep = "SEPARATOR ', '" if db_type == 'mysql' else ", ', '"
        rows = db_fetchall(conn, db_type, f"""
            SELECT v.id, v.codigo, v.criado_em, v.metodo_pagamento, v.estado, v.total,
                   pc.nome AS cliente_nome, c.nif AS cliente_nif,
                   pf.nome AS funcionario_nome,
                   GROUP_CONCAT(vi.nome {sep}) AS produtos,
                   COUNT(iv.id) AS num_itens
            FROM vendas v
            LEFT JOIN clientes c     ON v.cliente_id = c.id
            LEFT JOIN pessoas pc     ON c.pessoa_id = pc.id
            LEFT JOIN funcionarios f ON v.funcionario_id = f.id
            LEFT JOIN pessoas pf     ON f.pessoa_id = pf.id
            LEFT JOIN itens_venda iv ON iv.venda_id = v.id
            LEFT JOIN vinhos vi      ON iv.vinho_id = vi.id
            GROUP BY v.id ORDER BY v.criado_em DESC
        """)
        conn.close()

        wb = Workbook()
        ws = wb.active
        ws.title = 'Vendas'

        _add_brand_header(ws, 'Histórico de Vendas', 'Listagem completa de transações')

        headers = ['Código', 'Data / Hora', 'Cliente', 'NIF', 'Funcionário', 'Itens', 'Nº Itens', 'Método', 'Estado', 'Total (€)']
        header_row = 5
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=header_row, column=col_idx, value=h)
        _fmt_header_row(ws, header_row, headers)

        # Estilos de linha
        thin = Side(border_style='thin', color='ECE3D2')
        cell_border = Border(bottom=thin)
        ink_font = Font(name='Inter', size=10, color=_BRAND_INK)
        gold_font = Font(name='Cormorant Garamond', bold=True, size=11, color='8B7048')
        zebra = PatternFill('solid', fgColor='FAF6EC')

        total_geral = 0.0
        for i, r in enumerate(rows):
            row_i = header_row + 1 + i
            data_str = str(r['criado_em']) if r.get('criado_em') else ''
            try:
                dt = datetime.datetime.strptime(data_str.split('.')[0], '%Y-%m-%d %H:%M:%S')
                data_fmt = dt.strftime('%d/%m/%Y %H:%M')
            except Exception:
                data_fmt = data_str
            total = float(r.get('total') or 0)
            total_geral += total
            values = [
                r.get('codigo') or f"VD-{r['id']:04d}",
                data_fmt,
                r.get('cliente_nome') or 'Consumidor Final',
                r.get('cliente_nif') or '',
                r.get('funcionario_nome') or 'Sistema',
                r.get('produtos') or '',
                int(r.get('num_itens') or 0),
                r.get('metodo_pagamento') or '',
                r.get('estado') or '',
                total,
            ]
            for col_idx, v in enumerate(values, 1):
                c = ws.cell(row=row_i, column=col_idx, value=v)
                c.font = gold_font if col_idx == 10 else ink_font
                c.border = cell_border
                if col_idx == 10:
                    c.number_format = '#,##0.00 €'
                    c.alignment = Alignment(horizontal='right')
                if i % 2 == 1:
                    c.fill = zebra
            ws.row_dimensions[row_i].height = 22

        # Linha de totais
        totals_row = header_row + 1 + len(rows) + 1
        ws.cell(row=totals_row, column=9, value='TOTAL').font = Font(name='Inter', bold=True, size=11, color=_BRAND_INK)
        ws.cell(row=totals_row, column=9).alignment = Alignment(horizontal='right')
        tot_cell = ws.cell(row=totals_row, column=10, value=round(total_geral, 2))
        tot_cell.font = Font(name='Cormorant Garamond', bold=True, size=14, color=_BRAND_GOLD)
        tot_cell.number_format = '#,##0.00 €'
        tot_cell.alignment = Alignment(horizontal='right')
        ws.cell(row=totals_row, column=10).border = Border(top=Side(border_style='medium', color=_BRAND_INK))

        # Larguras de coluna
        widths = [16, 18, 28, 14, 24, 50, 10, 14, 14, 14]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
        ws.sheet_view.showGridLines = False

        # Output
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        fname = f"the100s-vendas-{datetime.datetime.now().strftime('%Y%m%d-%H%M')}.xlsx"
        return send_file(bio,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=fname)
    except Exception as e:
        log.error("Erro export vendas xlsx: %s", e)
        return jsonify({'error': str(e)}), 500


@app.route('/api/export/relatorio.xlsx', methods=['GET'])
def export_relatorio_xlsx():
    """Exporta relatório de gestão completo (KPIs + vendas mensais + top vinhos + stock crítico)."""
    if not OPENPYXL_AVAILABLE:
        return jsonify({'error': 'openpyxl não está instalado'}), 503
    try:
        conn, db_type = get_db()

        # KPIs
        receita = (db_fetchone(conn, db_type, "SELECT COALESCE(SUM(total),0) v FROM vendas WHERE estado='CONCLUIDA'") or {}).get('v', 0)
        n_vendas = (db_fetchone(conn, db_type, "SELECT COUNT(*) v FROM vendas WHERE estado='CONCLUIDA'") or {}).get('v', 0)
        ticket = round(float(receita) / n_vendas, 2) if n_vendas else 0
        valor_stock = (db_fetchone(conn, db_type, "SELECT COALESCE(SUM(preco*quantidade),0) v FROM vinhos") or {}).get('v', 0)
        n_vinhos = (db_fetchone(conn, db_type, "SELECT COUNT(*) v FROM vinhos") or {}).get('v', 0)
        n_critico = (db_fetchone(conn, db_type, "SELECT COUNT(*) v FROM vinhos WHERE quantidade<10") or {}).get('v', 0)

        # Vendas mensais — % precisa de ser escapado em MySQL (%% → %) porque o cursor faz format-string
        if db_type == 'mysql':
            mensais = db_fetchall(conn, db_type, """
                SELECT DATE_FORMAT(criado_em,'%%Y-%%m') mes, COUNT(*) n, SUM(total) tot, AVG(total) tk
                FROM vendas WHERE estado='CONCLUIDA'
                GROUP BY DATE_FORMAT(criado_em,'%%Y-%%m') ORDER BY mes DESC LIMIT 12
            """)
        else:
            mensais = db_fetchall(conn, db_type, """
                SELECT strftime('%Y-%m', criado_em) mes, COUNT(*) n, SUM(total) tot, AVG(total) tk
                FROM vendas WHERE estado='CONCLUIDA'
                GROUP BY strftime('%Y-%m', criado_em) ORDER BY mes DESC LIMIT 12
            """)

        # Top vinhos
        top = db_fetchall(conn, db_type, """
            SELECT v.nome, v.tipo, v.regiao, v.produtor,
                   COALESCE(SUM(iv.quantidade),0) qtd,
                   COALESCE(SUM(iv.quantidade*iv.preco_unit),0) receita,
                   v.preco
            FROM vinhos v LEFT JOIN itens_venda iv ON v.id = iv.vinho_id
            GROUP BY v.id ORDER BY receita DESC LIMIT 20
        """)

        # Stock crítico
        criticos = db_fetchall(conn, db_type,
            "SELECT nome, tipo, regiao, produtor, quantidade, preco FROM vinhos WHERE quantidade<10 ORDER BY quantidade ASC")
        conn.close()

        wb = Workbook()
        # ───── Sheet 1: KPIs ─────
        ws1 = wb.active
        ws1.title = 'Resumo'
        _add_brand_header(ws1, 'Relatório de Gestão', 'Snapshot KPIs · Vendas Mensais · Top Vinhos · Stock Crítico')

        ws1.cell(row=5, column=1, value='INDICADORES-CHAVE').font = Font(name='Inter', bold=True, size=11, color=_BRAND_GOLD)
        kpi_pairs = [
            ('Receita total (vendas concluídas)', float(receita), '€'),
            ('Nº de vendas', n_vendas, ''),
            ('Ticket médio', ticket, '€'),
            ('Valor de inventário', float(valor_stock), '€'),
            ('Total de referências', n_vinhos, ''),
            ('Linhas de stock críticas (<10)', n_critico, ''),
        ]
        ink_font = Font(name='Inter', size=10, color=_BRAND_INK)
        muted_font = Font(name='Inter', size=10, color=_BRAND_MUTED)
        gold_big = Font(name='Cormorant Garamond', bold=True, size=14, color=_BRAND_GOLD)
        for i, (lbl, val, sym) in enumerate(kpi_pairs):
            row_i = 6 + i
            c1 = ws1.cell(row=row_i, column=1, value=lbl); c1.font = muted_font
            c2 = ws1.cell(row=row_i, column=2, value=val); c2.font = gold_big
            c2.alignment = Alignment(horizontal='right')
            if sym == '€':
                c2.number_format = '#,##0.00 €'
            else:
                c2.number_format = '#,##0'
        ws1.column_dimensions['A'].width = 38
        ws1.column_dimensions['B'].width = 22
        ws1.sheet_view.showGridLines = False

        # ───── Sheet 2: Vendas Mensais ─────
        ws2 = wb.create_sheet('Vendas Mensais')
        _add_brand_header(ws2, 'Vendas por Mês', 'Últimos 12 meses')
        headers2 = ['Mês', 'Nº Vendas', 'Total (€)', 'Ticket Médio (€)']
        for i, h in enumerate(headers2, 1):
            ws2.cell(row=5, column=i, value=h)
        _fmt_header_row(ws2, 5, headers2)
        for i, m in enumerate(mensais):
            r = 6 + i
            ws2.cell(row=r, column=1, value=m['mes']).font = ink_font
            ws2.cell(row=r, column=2, value=int(m['n'] or 0)).font = ink_font
            t = ws2.cell(row=r, column=3, value=float(m['tot'] or 0)); t.font = Font(name='Cormorant Garamond', bold=True, size=11, color='8B7048'); t.number_format = '#,##0.00 €'
            tk = ws2.cell(row=r, column=4, value=float(m['tk'] or 0)); tk.font = ink_font; tk.number_format = '#,##0.00 €'
        for i, w in enumerate([14, 14, 16, 18], 1):
            ws2.column_dimensions[get_column_letter(i)].width = w
        ws2.sheet_view.showGridLines = False

        # ───── Sheet 3: Top Vinhos ─────
        ws3 = wb.create_sheet('Top Vinhos')
        _add_brand_header(ws3, 'Ranking de Produtos', 'Top 20 por receita acumulada')
        headers3 = ['Posição', 'Vinho', 'Tipo', 'Região', 'Produtor', 'Qtd. Vendida', 'PVP (€)', 'Receita (€)']
        for i, h in enumerate(headers3, 1):
            ws3.cell(row=5, column=i, value=h)
        _fmt_header_row(ws3, 5, headers3)
        for i, t in enumerate(top):
            r = 6 + i
            ws3.cell(row=r, column=1, value=i + 1).font = Font(name='Cormorant Garamond', bold=True, size=12, color=_BRAND_GOLD)
            ws3.cell(row=r, column=2, value=t['nome']).font = Font(name='Inter', bold=True, size=10, color=_BRAND_INK)
            ws3.cell(row=r, column=3, value=t.get('tipo') or '').font = ink_font
            ws3.cell(row=r, column=4, value=t.get('regiao') or '').font = ink_font
            ws3.cell(row=r, column=5, value=t.get('produtor') or '').font = ink_font
            ws3.cell(row=r, column=6, value=int(t.get('qtd') or 0)).font = ink_font
            p = ws3.cell(row=r, column=7, value=float(t.get('preco') or 0)); p.font = ink_font; p.number_format = '#,##0.00 €'
            rec = ws3.cell(row=r, column=8, value=float(t.get('receita') or 0)); rec.font = Font(name='Cormorant Garamond', bold=True, size=11, color='8B7048'); rec.number_format = '#,##0.00 €'
        for i, w in enumerate([10, 32, 12, 16, 22, 14, 12, 16], 1):
            ws3.column_dimensions[get_column_letter(i)].width = w
        ws3.sheet_view.showGridLines = False

        # ───── Sheet 4: Stock Crítico ─────
        ws4 = wb.create_sheet('Stock Crítico')
        _add_brand_header(ws4, 'Stock Crítico', 'Referências com menos de 10 unidades')
        headers4 = ['Vinho', 'Tipo', 'Região', 'Produtor', 'Stock', 'PVP (€)']
        for i, h in enumerate(headers4, 1):
            ws4.cell(row=5, column=i, value=h)
        _fmt_header_row(ws4, 5, headers4, fill_hex='8B1A14', font_color='F5F0E8')
        red_font = Font(name='Inter', bold=True, size=10, color='8B1A14')
        for i, c in enumerate(criticos):
            r = 6 + i
            ws4.cell(row=r, column=1, value=c['nome']).font = Font(name='Inter', bold=True, size=10)
            ws4.cell(row=r, column=2, value=c.get('tipo') or '').font = ink_font
            ws4.cell(row=r, column=3, value=c.get('regiao') or '').font = ink_font
            ws4.cell(row=r, column=4, value=c.get('produtor') or '').font = ink_font
            qty_cell = ws4.cell(row=r, column=5, value=int(c.get('quantidade') or 0))
            qty_cell.font = red_font
            qty_cell.alignment = Alignment(horizontal='right')
            p = ws4.cell(row=r, column=6, value=float(c.get('preco') or 0)); p.font = ink_font; p.number_format = '#,##0.00 €'
        for i, w in enumerate([34, 12, 16, 22, 10, 14], 1):
            ws4.column_dimensions[get_column_letter(i)].width = w
        ws4.sheet_view.showGridLines = False

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        fname = f"the100s-relatorio-{datetime.datetime.now().strftime('%Y%m%d-%H%M')}.xlsx"
        return send_file(bio,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=fname)
    except Exception as e:
        log.error("Erro export relatorio xlsx: %s", e)
        return jsonify({'error': str(e)}), 500

# ── MAIN ──────────────────────────────────────────────────
if __name__ == '__main__':
    # Testar conexão MySQL
    if MYSQL_AVAILABLE:
        try:
            cfg = {**MYSQL_CFG}
            cfg.pop('cursorclass', None)
            cfg['cursorclass'] = pymysql.cursors.DictCursor
            test = pymysql.connect(**cfg)
            test.close()
            log.info("MySQL ligado com sucesso")
        except Exception as e:
            log.warning("MySQL não disponível: %s", e)
            log.info("A usar SQLite como fallback")
            DB_TYPE = 'sqlite'
    else:
        log.info("PyMySQL não instalado, a usar SQLite")

    init_db()

    db_label = "MySQL (vinhadouro)" if DB_TYPE == 'mysql' else "SQLite (vinhadouro.db)"
    print("\n" + "="*55)
    print("  the 100's — Sistema Premium v3.2")
    print("="*55)
    print(f"  Site:     http://localhost:8080")
    print(f"  API:      http://localhost:8080/api/")
    print(f"  BD:       {db_label}")
    print("="*55)
    print("  Endpoints:")
    print("    /api/vinhos            - Gestao de vinhos")
    print("    /api/vendas            - Vendas e devoluções")
    print("    /api/caves             - Caves de armazenamento")
    print("    /api/movimentos-stock  - Historico de stock")
    print("    /api/funcionarios      - Gestao de equipa")
    print("    /api/clientes          - Clientes da loja")
    print("    /api/relatorios/*      - Relatorios")
    print("    /api/dashboard         - Dashboard do gerente")
    print("="*55)
    print("  Abre http://localhost:8080 no browser!\n")
    app.run(host='0.0.0.0', port=8080, debug=False)
